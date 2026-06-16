# ── PASSO C (D+1): escreve "Odd Fech." na planilha a partir do CSV ─────
# Baixe o fechamentos.csv do repo (atualizado pelo Actions) e rode isto.
import csv
from openpyxl import load_workbook

ARQUIVO     = "picks_2026-06-15.xlsx"
FECHAMENTOS = "fechamentos.csv"

with open(FECHAMENTOS, newline="", encoding="utf-8") as f:
    fech = {f"{r['partida']}|{r['mercado']}|{r['aposta']}": float(r["odd_fech"])
            for r in csv.DictReader(f)}

wb = load_workbook(ARQUIVO); ws = wb["Avaliações"]
H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
n = 0
for r in range(2, ws.max_row + 1):
    if ws.cell(r, H["Odd Fech."]).value not in (None, ""): continue
    k = (f"{ws.cell(r,H['Partida']).value}|{ws.cell(r,H['Mercado']).value}|"
         f"{ws.cell(r,H['Aposta']).value}")
    if k in fech:
        ws.cell(r, H["Odd Fech."]).value = fech[k]; n += 1
wb.save(ARQUIVO)
print(f"✅ {n} linha(s) com 'Odd Fech.' preenchida. CLV % calcula sozinha ao abrir.")
