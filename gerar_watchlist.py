# ── PASSO A (D0): exporta os picks pendentes p/ watchlist.csv ──────────
# Rode depois do motor. Sobe o watchlist.csv gerado no seu repo do GitHub.
import csv
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook

ARQUIVO = "picks_2026-06-15.xlsx"   # planilha do dia
ANO     = 2026
SO_SUGERIDOS = True                 # True = só Sugerido=Sim
BRT = timezone(timedelta(hours=-3))

wb = load_workbook(ARQUIVO); ws = wb["Avaliações"]
H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
linhas = []
for r in range(2, ws.max_row + 1):
    if ws.cell(r, H["Odd Fech."]).value not in (None, ""):      # já tem fechamento
        continue
    if SO_SUGERIDOS and "Sugerido" in H and ws.cell(r, H["Sugerido"]).value != "Sim":
        continue
    d, hhmm = ws.cell(r, H["Data"]).value, ws.cell(r, H["Hora"]).value
    liga, part = ws.cell(r, H["Liga"]).value, ws.cell(r, H["Partida"]).value
    merc, ap = ws.cell(r, H["Mercado"]).value, ws.cell(r, H["Aposta"]).value
    if not (d and hhmm and part): continue
    dd, mm = str(d).split("/"); hh, mi = str(hhmm).split(":")
    ko = datetime(int(ANO), int(mm), int(dd), int(hh), int(mi), tzinfo=BRT).astimezone(timezone.utc)
    linhas.append({"kickoff_iso": ko.isoformat().replace("+00:00", "Z"),
                   "liga": liga, "partida": part, "mercado": merc, "aposta": ap})

with open("watchlist.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["kickoff_iso","liga","partida","mercado","aposta"])
    w.writeheader(); w.writerows(linhas)
print(f"✅ watchlist.csv com {len(linhas)} pick(s) pendente(s). Suba no repo do GitHub.")
